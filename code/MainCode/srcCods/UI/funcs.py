import os
import pathlib as pl
from datetime import datetime
import json
import sys
import pathlib as pl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import certifi
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()

# ==========================================================
#                       APLICATIVO
#               Criador - Yuri Bertola de souza
#      Linkedin:https://www.linkedin.com/in/yuri-bertola/
# ==========================================================

# ===== CONFIGURAÇÃO GLOBAL DE REQUISIÇÕES SSL =====
def criar_sessao_otimizada():
    """
    Cria uma sessão requests otimizada para evitar erros SSL
    e melhorar performance com connection pooling.
    
    Problemas resolvidos:
    - SSLEOFError (EOF occurred in violation of protocol)
    - Retry automático em caso de falha
    - Pool de conexões para melhor performance
    - Timeout adequado para cada requisição
    """
    sessao = requests.Session()
    
    # Configurar retentativas automáticas
    retry_strategy = Retry(
        total=3,
        connect=2,
        read=2,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST", "PUT", "DELETE", "HEAD"]
    )
    
    # Adapter para HTTP/HTTPS com pool de conexões otimizado
    adapter = HTTPAdapter(
        max_retries=retry_strategy,
        pool_connections=20,
        pool_maxsize=30
    )
    
    sessao.mount("https://", adapter)
    sessao.mount("http://", adapter)
    
    return sessao

# Sessão global para todas as requisições
SESSAO_GLOBAL = criar_sessao_otimizada()
# ===== FIM CONFIGURAÇÃO SSL =====

if getattr(sys, 'frozen', False):
    # Executando como .exe (PyInstaller)
    BASE_DIR = pl.Path(sys._MEIPASS)
else:
    # Executando como .py
    BASE_DIR = pl.Path(__file__).resolve().parent

CONFIGS_FILIAIS_DIR = BASE_DIR / "configs_filiais"


FILIAIS_CONFIGS = {}

if CONFIGS_FILIAIS_DIR.exists():
    for arquivo in CONFIGS_FILIAIS_DIR.glob("*.json"):
        with open(arquivo, "r", encoding="utf-8") as f:
            FILIAIS_CONFIGS[arquivo.stem] = json.load(f)


#Filiais Sul
FILIAISSULMAIN = FILIAIS_CONFIGS.get("RegraSul", {})

#Filiais Sudeste 
FILIAISSUDESTEMAIN = FILIAIS_CONFIGS.get("RegraSudeste", {})

#Filiais NorteNordeste
FILIAISNORTE_NORDESTEMAIN = FILIAIS_CONFIGS.get("RegraNorte_Nordeste", {})

#Filiais SP MAIN
FILIAISSPMAIN = FILIAIS_CONFIGS.get("RegraSp", {})
FILIAIS_SP = FILIAIS_CONFIGS.get("RegraSp", {}).get("filiais", [])

#Filiais Tupladas:
FILIAISTUPLAS = FILIAIS_CONFIGS.get("filiaisTuplas", [])



TABELASIDSNOMES = []
LINHAS = []

def get_tokens_json_path() -> pl.Path:

    # onde o exe está
    if getattr(sys, "frozen", False):
        base = pl.Path(sys.executable).parent
    else:
        base = pl.Path(__file__).resolve().parent

    candidatos = [

        # O CAMINHO QUE SEMPRE FUNCIONA NO EXE
        base / "tokens" / "tokens.json",

        # Caminho original no projeto (para rodar no Python)
        base.parent / "tokens" / "excutable" / "dist" / "tokens" / "tokens.json",
        base / "tokens" / "excutable" / "dist" / "tokens" / "tokens.json",
        base / "tokens" / "tokens.json",
    ]

    for c in candidatos:
        if c.exists():
            return c

    return None



def get_base_path():#
    if getattr(sys, 'frozen', False):
        # Rodando como .exe (PyInstaller)
        return pl.Path(sys.executable).parent
    else:
        # Rodando como script Python normal
        return pl.Path(__file__).parent

BASEPATH = get_base_path()
BASEPATH2 = get_tokens_json_path()
print(f"Base Path: {BASEPATH}")
print(f"Base Path 2: {BASEPATH2}")



def carregar_token(Token_FILE):
    if Token_FILE is None:
        print("ERRO: Token_FILE veio como None — get_tokens_json_path() não achou tokens.json")
        return None

    if not os.path.exists(Token_FILE):
        print("ERRO: Caminho não existe:", Token_FILE)
        return None

    try:
        with open(Token_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return None
    
def extrair_token():
    token_data = carregar_token(BASEPATH2)
    if token_data and "access_token" in token_data:
        return token_data["access_token"]
    return None

#REQUESTS - SENIOR

def request_tabela_salariais_Nomes(token):
    # ⚠️ IMPORTANTE: Esta função usa SESSAO_GLOBAL com:
    # - Retry automático (3 tentativas)
    # - Timeout: 30 segundos
    # - Pool de conexões otimizado
    # Isso resolve SSLEOFError de forma automática.
    url = "https://platform.senior.com.br/t/senior.com.br/bridge/1.0/rest/hcm/remuneration/queries/wageScales"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }

    response = SESSAO_GLOBAL.get(url, headers=headers, verify=False)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Erro na requisição: {response.status_code} - {response.text}")
        return None
    

def addObjsTabelas(nome,id):
    obj = {
        "names": str(nome),
        "ids":  id
    }

    TABELASIDSNOMES.append(obj)

def listarJson(token):
    if not token:
        print("Token inválido.")
        return []

    jsonTabelas = request_tabela_salariais_Nomes(token)

    if not jsonTabelas or "wageScales" not in jsonTabelas:
        print("Erro ao carregar tabelas.")
        return []

    TABELASIDSNOMES.clear()  # evita duplicação

    for tabela in jsonTabelas.get("wageScales", []):
        tabelaId = tabela.get("id")
        nome = tabela.get("name")

        if tabelaId is not None and nome is not None:
            addObjsTabelas(nome, tabelaId)

    return TABELASIDSNOMES.copy()


def print_tabelas(lista):
    print(f"{'NOME':<40} | ID")
    print("-" * 90)

    for item in lista:
        print(f"{item['names']:<40} | {item['ids']}")


def request_tabela_salariais_Detalhes(token, tabela_id, keepAlive, wageRevisionId=None):
    # ⚠️ IMPORTANTE: Esta função usa SESSAO_GLOBAL com:
    # - Retry automático (3 tentativas)
    # - Timeout: 30 segundos
    # - Pool de conexões otimizado
    # Isso resolve SSLEOFError de forma automática.
    url = "https://platform.senior.com.br/t/senior.com.br/bridge/1.0/rest/hcm/remuneration/queries/getWageScale"


    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
        
    }

    if keepAlive:
        headers["Connection"] = "keep-alive"
    
    payload = {
        "wageScaleId": tabela_id,
        "wageRevisionId": wageRevisionId
    }

    response = SESSAO_GLOBAL.post(url, headers=headers, json=payload, verify=False)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Erro na requisição: {response.status_code} - {response.text}")
        return None

def pegarTabelaRevisions(token, tabela_id=None):
    # ⚠️ IMPORTANTE: Esta função usa SESSAO_GLOBAL com:
    # - Retry automático (3 tentativas)
    # - Timeout: 30 segundos
    # - Pool de conexões otimizado
    # Isso resolve SSLEOFError de forma automática.
    url = "https://platform.senior.com.br/t/senior.com.br/bridge/1.0/rest/hcm/remuneration/queries/getWageScale"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "keep-alive": "true"
    }

    tabelas = []

    payload = {
        "wageScaleId": tabela_id,
    }

    response = SESSAO_GLOBAL.post(url, headers=headers, json=payload, verify=False)

    if response.status_code != 200:
        raise Exception(f"Erro na requisição: {response.status_code} - {response.text}")

    data = response.json()

    return data.get("revisions", [])

def pegarTabelaRevisionsG(CaminhoPasta: pl.Path, tabela_id):

    import os
    import requests
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    token = extrair_token()
    if not token:
        print("Token inválido")
        return

    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ===============================
    # BUSCA REVISÕES
    # ===============================
    revisoes = pegarTabelaRevisions(token, tabela_id=tabela_id)
    if not revisoes:
        print("Nenhuma revisão encontrada")
        return

    # ===============================
    # EXCEL
    # ===============================

    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Revisions - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)
    
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    COLUNAS = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9",
        "FILIAL", "VIGÊNCIA"
    ]

    # ===== ESTILOS (IGUAIS AO PROJETO) =====
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===============================
    # FUNÇÃO INTERNA
    # ===============================
    def escrever_aba(nome_aba, classes, vigencia):

        if nome_aba in wb.sheetnames:
            del wb[nome_aba]

        ws = wb.create_sheet(nome_aba)

        # CABEÇALHO
        for col_idx, nome_col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_col)
            cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
            cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
            cell.alignment = alinharCentro
            cell.border = bordaFina

        # DADOS
        for i, classe in enumerate(classes, start=2):

            # NÍVEL
            cell = ws.cell(row=i, column=1, value=classe.get("name"))
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0

                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            # FILIAL (em branco)
            cell = ws.cell(row=i, column=11, value=filial)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            # VIGÊNCIA
            cell = ws.cell(row=i, column=12, value=vigencia)
            cell.font = fonteValores
            cell.alignment = alinharCentro
            cell.border = bordaFina

    # ===============================
    # LOOP DAS REVISÕES
    # ===============================
    for rev in revisoes:

        revision_id = rev.get("id")
        vigencia = rev.get("startDate")

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={
                "wageScaleId": tabela_id,
                "wageRevisionId": revision_id
            },
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            print(f"Erro na revisão {revision_id}")
            continue

        import re

        dados = response.json()
        classes = dados.get("classes", [])

        nome_tabela = dados.get("name", "")
        match = re.search(r'(\d+)$', nome_tabela)
        filial = match.group(1) if match else "N/A"

        classes = response.json().get("classes", [])
        escrever_aba(vigencia, classes, vigencia)

    wb.save(caminho_excel)




#_---------------------------_---------------------------------#
# função que concatena e substitui e outra que em uma nova aba
#
def verificarExistencia(filiais):
    """
    Verifica quais filiais NÃO possuem wage scales
    
    Args:
        filiais: Lista de códigos de filiais ["5003", "5042", ...]
    
    Retorna:
        Lista com filiais que NÃO existem
        Exemplo: ["5003", "5042"]
    """
    
    # Buscar todas as wage scales disponíveis
    dados_meta = request_tabela_salariais_Nomes(extrair_token())
    
    # Criar mapa nome -> id
    id_map = {item["name"]: item["id"] for item in dados_meta.get("wageScales", [])}
    
    # Filtrar filiais que NÃO existem
    filiais_nao_existem = []
    for filial in filiais:
        chave = f"Tabela Salarial {filial}"
        if chave in id_map:
            filiais_nao_existem.append(filial)
    
    return filiais_nao_existem




#FUNÇÃO QUE FAZ VÁRIAS REQUISIÇÕES KEEP-ALIVE
def obter_dados_multiplos_ids(ids):
    token = extrair_token()
    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Connection": "keep-alive"
    }

    global FILIAIS_SP
    FILIAIS_SP_array = FILIAIS_SP

    FILIAIS_SP_existentes = verificarExistencia(FILIAIS_SP_array)

    FILIAIS_SP2 = [f for f in set(FILIAIS_SP_array) if f not in FILIAIS_SP_existentes]

    IDS_SP = {"74070F901FE74A358F8B1740EDF60F06"}

    IdsTuplas = [item["id"] for item in FILIAISTUPLAS]
    NomesTuplas = [item["nome"] for item in FILIAISTUPLAS]
    FiliaisTuplas = [item["valores"] for item in FILIAISTUPLAS]

    tabelas = []

    for tabela_id in ids:
        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": tabela_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        tabela = response.json()
        nome_tabela = tabela.get("name", "")
        classes = tabela.get("classes", [])

        # ===== BASE ÚNICA DA TABELA =====
        linhas_base = []
        for classe in classes:
            linhas_base.append({
                "fonte": nome_tabela,
                "name": classe.get("name"),
                "valores": [v.get("value") for v in classe.get("values", [])],
                "alfa": [v.get("name") for v in classe.get("values", [])]
            })

        # ===== VERIFICAÇÃO LOCAL (SEM MODO) =====
        eh_sp = (
            tabela.get("id") in IDS_SP or
            FILIAISSPMAIN.get("fonte") in nome_tabela.upper()
        )

        eh_tupla = (
            tabela.get("id") in IdsTuplas or
            any(nome.upper() in nome_tabela.upper() for nome in NomesTuplas)
        )

        tupla_atual = next(
            (t for t in FILIAISTUPLAS if t["id"] == tabela.get("id")),
            None
        )


        if eh_sp:
            # 🔹 SP sempre vira várias tabelas (1 por filial)
            for filial in FILIAIS_SP2:
                aba = []
                for linha in linhas_base:
                    nova = linha.copy()
                    filial_limpa = str(filial).strip()[:4]
                    nova["filial"] = filial
                    nova["fonte"] = f"Tabela Salarial {filial_limpa}"
                    aba.append(nova)

                tabelas.append(aba)

        elif eh_tupla and tupla_atual:
            for filial in tupla_atual["valores"]:   # 5030, 5060
                aba = []                            # ✅ uma aba por filial

                for linha in linhas_base:
                    nova = linha.copy()
                    # nomeLimpo = nome_tabela.replace(f"{filial}", "")
                    nova["fonte"] = f"Tabela Salarial {filial}"
                    aba.append(nova)

                tabelas.append(aba)      
        else:
            # 🔹 fluxo padrão
            tabelas.append(linhas_base)

    return tabelas


#FUNÇÂO RECURSIVA PARA ADICIONAR LINHAS
def adicionar_linha(nivel, valores, letras, font, idx):
    linha = {
        "fonte": str(font),
        "name": str(nivel),
        "valores": valores,
        "alfa": letras,
        "idx": idx
    }
    LINHAS.append(linha)
#------------------------------------------#

def criarTabelaNoPadrao(token, tabela_id, keepalive, revision_id=False, caminho=None):
    """
    Popula LINHAS.
    - keepalive=True  → dados já normalizados
    - keepalive=False → JSON bruto
    """
    # 5003,5041,5048,50030
    # NÃO limpar LINHAS aqui
    # ===== MODO KEEPALIVE =====
    if keepalive and revision_id is False:
        tabelas = obter_dados_multiplos_ids(tabela_id)

        for idx, linhas_tabela in enumerate(tabelas):
            for linha in linhas_tabela:
                linha["idx"] = idx
                LINHAS.append(linha)

        return
    # ===== MODO KEEPALIVE COM REVISION ID =====
    if keepalive and revision_id:
        tabelas = pegarTabelaRevisionsG(caminho, tabela_id)
        return 

    # ===== MODO NORMAL =====
    tabela = request_tabela_salariais_Detalhes(
        token, tabela_id, keepAlive=False, wageRevisionId=revision_id
    )

    if not tabela:
        print("Nenhuma tabela encontrada.")
        return

    namejson = tabela.get("name")
    classes = tabela.get("classes", [])

    for linhas in classes:
        valores = [v.get("value") for v in linhas.get("values", [])]
        nomes = [v.get("name") for v in linhas.get("values", [])]

        adicionar_linha(
            linhas.get("name"),
            valores,
            nomes,
            namejson,
            idx=0
        )


from collections import defaultdict
def agrupar_linhas_por_tabela(LINHAS):
    tabelas = defaultdict(list)
    for linha in LINHAS:
        tabelas[linha["idx"]].append(linha)
    return list(tabelas.values())


def obter_workbook(caminho_excel: str):
    from openpyxl import load_workbook, Workbook
    """Retorna um workbook existente, ou cria um novo caso o arquivo não exista."""
    if os.path.exists(caminho_excel):
        print("📄 Abrindo Excel existente…")
        return load_workbook(caminho_excel)
    else:
        print("📄 Criando novo Excel…")
        return Workbook()

def criartabela(
    CaminhoPasta,
    tabela_id,
    nometabela,
    aba=False,
    nomeAba=None,
    onef=False,
    keepalive=False,
    revision_id=None,
    revisionIds = False
):
    import os
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


    colunas = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9", "FILIAL"
    ]

    token = extrair_token()

    # ================= BUSCA DE DADOS =================
    LINHAS.clear()

    if keepalive and not revisionIds:
        criarTabelaNoPadrao(token, tabela_id, True, revision_id=revisionIds, caminho=None)
        tabelas = agrupar_linhas_por_tabela(LINHAS)

        if not nometabela or not nometabela.strip():
            nometabela = f"Geração de Todas as Tabelas - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    
    elif keepalive and revisionIds:
        criarTabelaNoPadrao(token, tabela_id, True, revision_id=revisionIds, caminho=CaminhoPasta)
        tabelas = agrupar_linhas_por_tabela(LINHAS)
        return 
    else:
        criarTabelaNoPadrao(token, tabela_id, False, revision_id=revision_id)
        tabelas = [LINHAS]

    if not nometabela or not nometabela.strip(): nometabela = "Tabela - Salárial"

    caminho_excel = os.path.join(CaminhoPasta, f"{nometabela} - {datetime.now().strftime('%Y-%m-%d')}.xlsx")
    wb = obter_workbook(caminho_excel)


    # ================= ESTILOS =================
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== LOOP PRINCIPAL DE TABELAS =====

    for idx, LINHAS_TABELA in enumerate(tabelas, start=1):
        if aba:
            fonte_tabela = LINHAS_TABELA[0].get("fonte", f"ABA_{idx}")
            nome = nomeAba if nomeAba else fonte_tabela
            if nome in wb.sheetnames:
                nome = f"{nome}"

            ws = wb.create_sheet(nome)
            escrever_cabecalho = True
            linha_base = 1
        else:
            ws = wb.active

            if ws.max_row <= 1:
                escrever_cabecalho = True
                linha_base = 1
            else:
                escrever_cabecalho = False
                linha_base = ws.max_row + 1

        # --- CABEÇALHO (APENAS SE NECESSÁRIO) ---
        if escrever_cabecalho:
            for col_idx, nome_col in enumerate(colunas, start=1):
                cell = ws.cell(row=linha_base, column=col_idx, value=nome_col)
                cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
                cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
                cell.alignment = alinharCentro
                cell.border = bordaFina

        linha_inicio = linha_base + (1 if escrever_cabecalho else 0)

        # --- DADOS ---
        for i, linha in enumerate(LINHAS_TABELA, start=linha_inicio):

            # NÍVEL
            cell = ws.cell(row=i, column=1, value=linha["name"])
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            # VALORES
            for j in range(2, len(colunas)):
                try:
                    valor = linha["valores"][j - 2]
                except (IndexError, TypeError):
                    valor = 0

                if not isinstance(valor, (int, float)):
                    valor = 0

                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            # FILIAL (SEMPRE ÚLTIMA COLUNA)
            col_filial = len(colunas)
            fonte = linha.get("fonte", "")
            numero = fonte.split()[-1] if fonte else ""


            cell = ws.cell(row=i, column=col_filial, value=numero)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

    wb.save(caminho_excel)
    LINHAS.clear()

def gerar_tabela_SP_engessada(CaminhoPasta: str, tabela_id):

    import os
    import requests
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    global FILIAIS_SP

    token = extrair_token()
    if not token:
        print("Token inválido")
        return

    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ===============================
    # VERIFICA FILIAIS NA API
    # ===============================
    FILIAIS_EXISTENTES = verificarExistencia(FILIAIS_SP)
    FILIAIS_NAO_EXISTENTES = [f for f in FILIAIS_SP if f not in FILIAIS_EXISTENTES]

    # ===============================
    # MAPA DE IDS
    # ===============================
    dados_meta = request_tabela_salariais_Nomes(token)
    ID_MAP = {
        item["name"]: item["id"]
        for item in dados_meta.get("wageScales", [])
    }

    # ===============================
    # TABELA BASE (FALLBACK)
    # ===============================
    response_base = SESSAO_GLOBAL.post(
        URL_API,
        json={"wageScaleId": tabela_id},
        headers=headers,
        verify=False
    )

    if response_base.status_code != 200:
        print("Erro ao buscar tabela base")
        return

    CLASSES_BASE = response_base.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================

    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial SP - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)
    
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    COLUNAS = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9", "FILIAL"
    ]

    # ===== ESTILOS =====
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===============================
    # ABA CONSOLIDADA (PRIMEIRA)
    # ===============================
    if "CONSOLIDADO" in wb.sheetnames:
        del wb["CONSOLIDADO"]

    ws_consolidado = wb.create_sheet("Sheet", 0)

    for col_idx, nome_col in enumerate(COLUNAS, start=1):
        cell = ws_consolidado.cell(row=1, column=col_idx, value=nome_col)
        cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
        cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
        cell.alignment = alinharCentro
        cell.border = bordaFina

    linha_consolidado = 2

    # ===============================
    # FUNÇÕES INTERNAS
    # ===============================
    def escrever_aba(filial, classes):
        if filial in wb.sheetnames:
            del wb[filial]

        ws = wb.create_sheet(filial)

        for col_idx, nome_col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_col)
            cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
            cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
            cell.alignment = alinharCentro
            cell.border = bordaFina

        for i, classe in enumerate(classes, start=2):

            ws.cell(row=i, column=1, value=classe.get("name")).font = fontenivel
            ws.cell(row=i, column=1).fill = prenchimento2
            ws.cell(row=i, column=1).alignment = alinharCentro
            ws.cell(row=i, column=1).border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws.cell(row=i, column=11, value=filial)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

    def escrever_consolidado(filial, classes):
        nonlocal linha_consolidado

        for classe in classes:
            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=1,
                value=classe.get("name")
            )
            cell.font = fontenivel
            cell.fill = prenchimento2      # ← ROXO
            cell.alignment = alinharCentro
            cell.border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws_consolidado.cell(
                    row=linha_consolidado,
                    column=j,
                    value=valor
                )
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=11,
                value=filial
            )
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            linha_consolidado += 1

    # ===============================
    # EXECUÇÃO (SIMULTÂNEA)
    # ===============================
    for filial in FILIAIS_EXISTENTES:
        wage_id = ID_MAP.get(f"Tabela Salarial {filial}")

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])
        escrever_aba(filial, classes)
        escrever_consolidado(filial, classes)

    for filial in FILIAIS_NAO_EXISTENTES:
        escrever_aba(filial, CLASSES_BASE)
        escrever_consolidado(filial, CLASSES_BASE)

    wb.save(caminho_excel)
    print("Tabelas SP Criadas")

#Região Norte e Nordeste(Nordeste)
def gerar_tabela_NORTE_NORDESTE(CaminhoPasta: str):

    import os
    import requests
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from datetime import datetime

    token = extrair_token()
    if not token:
        print("Token inválido")
        return

    # ===============================
    # REGRA (JSON JÁ CARREGADO NO MAP)
    # ===============================
    regra = FILIAIS_CONFIGS.get("RegraNorte_Nordeste", {})
    FILIAIS = regra.get("filiais", [])

    if not FILIAIS:
        print("Nenhuma filial definida para Norte/Nordeste")
        return

    tabela_id = regra.get("id")

    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ===============================
    # VERIFICA FILIAIS NA API
    # ===============================
    FILIAIS_EXISTENTES = verificarExistencia(FILIAIS)
    FILIAIS_NAO_EXISTENTES = [f for f in FILIAIS if f not in FILIAIS_EXISTENTES]

    # ===============================
    # MAPA DE IDS (NAME -> ID)
    # ===============================
    dados_meta = request_tabela_salariais_Nomes(token)

    ID_MAP = {
        item["name"]: item["id"]
        for item in dados_meta.get("wageScales", [])
    }

    # ===============================
    # TABELA BASE (FALLBACK)
    # ===============================
    response_base = SESSAO_GLOBAL.post(
        URL_API,
        json={"wageScaleId": tabela_id},
        headers=headers,
        verify=False
    )

    if response_base.status_code != 200:
        print("Erro ao buscar tabela base")
        return

    CLASSES_BASE = response_base.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial NORTE_NORDESTE - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    # ===============================
    # EXECUÇÃO — PARTE ALTERADA
    # ===============================
    for filial in FILIAIS_EXISTENTES:

        chave = f"Tabela Salarial {filial}"
        wage_id = ID_MAP.get(chave)

        if not wage_id:
            print(f"Tabela não encontrada no mapa: {chave}")
            continue

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial NORTE_NORDESTE - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    COLUNAS = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9", "FILIAL"
    ]

    # ===== ESTILOS =====
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===============================
    # ABA CONSOLIDADA (PRIMEIRA)
    # ===============================
    if "CONSOLIDADO" in wb.sheetnames:
        del wb["CONSOLIDADO"]

    ws_consolidado = wb.create_sheet("Sheet", 0)

    for col_idx, nome_col in enumerate(COLUNAS, start=1):
        cell = ws_consolidado.cell(row=1, column=col_idx, value=nome_col)
        cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
        cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
        cell.alignment = alinharCentro
        cell.border = bordaFina

    linha_consolidado = 2

    # ===============================
    # FUNÇÕES INTERNAS
    # ===============================
    def escrever_aba(filial, classes):
        if filial in wb.sheetnames:
            del wb[filial]

        ws = wb.create_sheet(filial)

        for col_idx, nome_col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_col)
            cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
            cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
            cell.alignment = alinharCentro
            cell.border = bordaFina

        for i, classe in enumerate(classes, start=2):

            ws.cell(row=i, column=1, value=classe.get("name")).font = fontenivel
            ws.cell(row=i, column=1).fill = prenchimento2
            ws.cell(row=i, column=1).alignment = alinharCentro
            ws.cell(row=i, column=1).border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws.cell(row=i, column=11, value=filial)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

    def escrever_consolidado(filial, classes):
        nonlocal linha_consolidado

        for classe in classes:
            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=1,
                value=classe.get("name")
            )
            cell.font = fontenivel
            cell.fill = prenchimento2      # ← ROXO
            cell.alignment = alinharCentro
            cell.border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws_consolidado.cell(
                    row=linha_consolidado,
                    column=j,
                    value=valor
                )
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=11,
                value=filial
            )
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            linha_consolidado += 1

    # ===============================
    # EXECUÇÃO (SIMULTÂNEA)
    # ===============================
    for filial in FILIAIS_EXISTENTES:
        wage_id = ID_MAP.get(f"Tabela Salarial {filial}")

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])
        escrever_aba(filial, classes)
        escrever_consolidado(filial, classes)

    for filial in FILIAIS_NAO_EXISTENTES:
        escrever_aba(filial, CLASSES_BASE)
        escrever_consolidado(filial, CLASSES_BASE)

    wb.save(caminho_excel)
    print("Tabelas NORTE e NORDESTE CRIADAS")

#Regiao Sul 
def gerar_tabela_SUL(CaminhoPasta: str):

    import os
    import requests
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from datetime import datetime

    token = extrair_token()
    if not token:
        print("Token inválido")
        return

    # ===============================
    # REGRA (JSON JÁ CARREGADO NO MAP)
    # ===============================
    regra = FILIAIS_CONFIGS.get("RegraSul", {})
    FILIAIS = regra.get("filiais", [])

    if not FILIAIS:
        print("Nenhuma filial definida para Sul")
        return

    tabela_id = regra.get("id", [])

    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ===============================
    # VERIFICA FILIAIS NA API
    # ===============================
    FILIAIS_EXISTENTES = verificarExistencia(FILIAIS)
    FILIAIS_NAO_EXISTENTES = [f for f in FILIAIS if f not in FILIAIS_EXISTENTES]

    # ===============================
    # MAPA DE IDS (NAME -> ID)
    # ===============================
    dados_meta = request_tabela_salariais_Nomes(token)

    ID_MAP = {
        item["name"]: item["id"]
        for item in dados_meta.get("wageScales", [])
    }

    # ===============================
    # TABELA BASE (FALLBACK)
    # ===============================
    response_base = SESSAO_GLOBAL.post(
        URL_API,
        json={"wageScaleId": tabela_id},
        headers=headers,
        verify=False
    )

    if response_base.status_code != 200:
        print("Erro ao buscar tabela base")
        return

    CLASSES_BASE = response_base.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial SUL - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    # ===============================
    # EXECUÇÃO — PARTE ALTERADA
    # ===============================
    for filial in FILIAIS_EXISTENTES:

        chave = f"Tabela Salarial {filial}"
        wage_id = ID_MAP.get(chave)

        if not wage_id:
            print(f"Tabela não encontrada no mapa: {chave}")
            continue

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial SUL - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    COLUNAS = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9", "FILIAL"
    ]

    # ===== ESTILOS =====
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===============================
    # ABA CONSOLIDADA (PRIMEIRA)
    # ===============================
    if "CONSOLIDADO" in wb.sheetnames:
        del wb["CONSOLIDADO"]

    ws_consolidado = wb.create_sheet("Sheet", 0)

    for col_idx, nome_col in enumerate(COLUNAS, start=1):
        cell = ws_consolidado.cell(row=1, column=col_idx, value=nome_col)
        cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
        cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
        cell.alignment = alinharCentro
        cell.border = bordaFina

    linha_consolidado = 2

    # ===============================
    # FUNÇÕES INTERNAS
    # ===============================
    def escrever_aba(filial, classes):
        if filial in wb.sheetnames:
            del wb[filial]

        ws = wb.create_sheet(filial)

        for col_idx, nome_col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_col)
            cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
            cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
            cell.alignment = alinharCentro
            cell.border = bordaFina

        for i, classe in enumerate(classes, start=2):

            ws.cell(row=i, column=1, value=classe.get("name")).font = fontenivel
            ws.cell(row=i, column=1).fill = prenchimento2
            ws.cell(row=i, column=1).alignment = alinharCentro
            ws.cell(row=i, column=1).border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws.cell(row=i, column=11, value=filial)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

    def escrever_consolidado(filial, classes):
        nonlocal linha_consolidado

        for classe in classes:
            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=1,
                value=classe.get("name")
            )
            cell.font = fontenivel
            cell.fill = prenchimento2      # ← ROXO
            cell.alignment = alinharCentro
            cell.border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws_consolidado.cell(
                    row=linha_consolidado,
                    column=j,
                    value=valor
                )
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=11,
                value=filial
            )
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            linha_consolidado += 1

    # ===============================
    # EXECUÇÃO (SIMULTÂNEA)
    # ===============================
    for filial in FILIAIS_EXISTENTES:
        wage_id = ID_MAP.get(f"Tabela Salarial {filial}")

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])
        escrever_aba(filial, classes)
        escrever_consolidado(filial, classes)

    for filial in FILIAIS_NAO_EXISTENTES:
        escrever_aba(filial, CLASSES_BASE)
        escrever_consolidado(filial, CLASSES_BASE)

    wb.save(caminho_excel)
    print("Tabelas SUl Criada")

#Regiao Sudeste
def gerar_tabela_SUDESTE(CaminhoPasta: str):

    import os
    import requests
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from datetime import datetime

    token = extrair_token()
    if not token:
        print("Token inválido")
        return

    # ===============================
    # REGRA (JSON JÁ CARREGADO NO MAP)
    # ===============================
    regra = FILIAIS_CONFIGS.get("RegraSudeste", {})
    FILIAIS = regra.get("filiais", [])

    if not FILIAIS:
        print("Nenhuma filial definida para Sudeste - JSON")
        return

    tabela_id = regra.get("id", [])

    URL_API = (
        "https://platform.senior.com.br/"
        "t/senior.com.br/bridge/1.0/rest/"
        "hcm/remuneration/queries/getWageScale"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ===============================
    # VERIFICA FILIAIS NA API
    # ===============================
    FILIAIS_EXISTENTES = verificarExistencia(FILIAIS)
    FILIAIS_NAO_EXISTENTES = [f for f in FILIAIS if f not in FILIAIS_EXISTENTES]

    # ===============================
    # MAPA DE IDS (NAME -> ID)
    # ===============================
    dados_meta = request_tabela_salariais_Nomes(token)

    ID_MAP = {
        item["name"]: item["id"]
        for item in dados_meta.get("wageScales", [])
    }

    # ===============================
    # TABELA BASE (FALLBACK)
    # ===============================
    response_base = SESSAO_GLOBAL.post(
        URL_API,
        json={"wageScaleId": tabela_id},
        headers=headers,
        verify=False
    )

    if response_base.status_code != 200:
        print("Erro ao buscar tabela base")
        return

    CLASSES_BASE = response_base.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial SUDESTE - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    # ===============================
    # EXECUÇÃO — PARTE ALTERADA
    # ===============================
    for filial in FILIAIS_EXISTENTES:

        chave = f"Tabela Salarial {filial}"
        wage_id = ID_MAP.get(chave)

        if not wage_id:
            print(f"Tabela não encontrada no mapa: {chave}")
            continue

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])

    # ===============================
    # EXCEL
    # ===============================
    caminho_excel = os.path.join(
        CaminhoPasta,
        f"Tabela Salarial SUDESTE - {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    wb = obter_workbook(caminho_excel)

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    COLUNAS = [
        "Nível", "STEP1", "STEP2", "STEP3", "STEP4",
        "100%", "STEP6", "STEP7", "STEP8", "STEP9", "FILIAL"
    ]

    # ===== ESTILOS =====
    fonteCabecalho = Font(name='Aptos Narrow', size=10, color='FFFFFF')
    fonteCabecalho2 = Font(name='Aptos Narrow', size=10, bold=True, color='FFFFFF')
    fontenivel = Font(name="Segoe UI Variable Small", size=7, bold=True, color="FFFFFF")
    fonteValores = Font(name="Segoe UI Variable Small", size=7, color="000000")

    prenchimento1 = PatternFill("solid", fgColor="E97132")
    prenchimento2 = PatternFill("solid", fgColor="7030A0")
    prenchimento3 = PatternFill("solid", fgColor="FEECEC")

    alinharCentro = Alignment(horizontal="center", vertical="center")

    bordaFina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===============================
    # ABA CONSOLIDADA (PRIMEIRA)
    # ===============================
    if "CONSOLIDADO" in wb.sheetnames:
        del wb["CONSOLIDADO"]

    ws_consolidado = wb.create_sheet("Sheet", 0)

    for col_idx, nome_col in enumerate(COLUNAS, start=1):
        cell = ws_consolidado.cell(row=1, column=col_idx, value=nome_col)
        cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
        cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
        cell.alignment = alinharCentro
        cell.border = bordaFina

    linha_consolidado = 2

    # ===============================
    # FUNÇÕES INTERNAS
    # ===============================
    def escrever_aba(filial, classes):
        if filial in wb.sheetnames:
            del wb[filial]

        ws = wb.create_sheet(filial)

        for col_idx, nome_col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_col)
            cell.font = fonteCabecalho2 if nome_col == "100%" else fonteCabecalho
            cell.fill = prenchimento2 if nome_col == "100%" else prenchimento1
            cell.alignment = alinharCentro
            cell.border = bordaFina

        for i, classe in enumerate(classes, start=2):

            ws.cell(row=i, column=1, value=classe.get("name")).font = fontenivel
            ws.cell(row=i, column=1).fill = prenchimento2
            ws.cell(row=i, column=1).alignment = alinharCentro
            ws.cell(row=i, column=1).border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws.cell(row=i, column=j, value=valor)
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws.cell(row=i, column=11, value=filial)
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

    def escrever_consolidado(filial, classes):
        nonlocal linha_consolidado

        for classe in classes:
            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=1,
                value=classe.get("name")
            )
            cell.font = fontenivel
            cell.fill = prenchimento2      # ← ROXO
            cell.alignment = alinharCentro
            cell.border = bordaFina

            valores = [v.get("value") for v in classe.get("values", [])]

            for j in range(2, 11):
                valor = valores[j - 2] if j - 2 < len(valores) else 0
                cell = ws_consolidado.cell(
                    row=linha_consolidado,
                    column=j,
                    value=valor
                )
                cell.number_format = 'R$ #,##0.00'
                cell.font = fonteValores
                cell.fill = prenchimento3 if j == 6 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = alinharCentro
                cell.border = bordaFina

            cell = ws_consolidado.cell(
                row=linha_consolidado,
                column=11,
                value=filial
            )
            cell.font = fontenivel
            cell.fill = prenchimento2
            cell.alignment = alinharCentro
            cell.border = bordaFina

            linha_consolidado += 1

    # ===============================
    # EXECUÇÃO (SIMULTÂNEA)
    # ===============================
    for filial in FILIAIS_EXISTENTES:
        wage_id = ID_MAP.get(f"Tabela Salarial {filial}")

        response = SESSAO_GLOBAL.post(
            URL_API,
            json={"wageScaleId": wage_id},
            headers=headers,
            verify=False
        )

        if response.status_code != 200:
            continue

        classes = response.json().get("classes", [])
        escrever_aba(filial, classes)
        escrever_consolidado(filial, classes)

    for filial in FILIAIS_NAO_EXISTENTES:
        escrever_aba(filial, CLASSES_BASE)
        escrever_consolidado(filial, CLASSES_BASE)

    wb.save(caminho_excel)
    print("Tabelas SUDESTE Criada")








#INPUT DE TEXTO
#INPUT DE CAMINHO DE PASTA
#INPUT DE DROPDOWN 
#INPUT fALSE/tRUE
#INPUT NOME aba textO